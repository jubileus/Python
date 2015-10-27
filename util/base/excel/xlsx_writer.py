# coding=UTF-8

from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
import openpyxl

__author__ = 'jubileus'

'''
    此工具类用于写入07+版excel
'''

# 创建要写入数据的workbook
def create_workbook():
    openpyxl.Workbook.encoding = 'gbk'
    return Workbook()


# 创建Writer
def create_writer(wb):
    return ExcelWriter(workbook=wb)


# 创建sheet
def create_sheet(wb, ws_name='Sheet1', ws_index=0):
    ws = wb.worksheets[ws_index]
    ws.title = ws_name
    return ws


# 保存workbook到磁盘
def save_workbook(ew, path='excel_1.xls'):
    ew.save(filename=path)


# 写入数据
def write(data, ws):
    for r in range(len(data)):
        row = data[r]
        for c in range(len(row)):
            col = get_column_letter(c + 1)
            ws.cell('{col}{row}'.format(col=col, row=r + 1)).value = row[c]