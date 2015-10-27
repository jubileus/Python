# coding=UTF-8

import traceback
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

__author__ = 'jubileus'

'''
   此工具类用于07+excel的读取
'''


# 加要读取的workbook
def get_workbook(path, read_only=True):
    try:
        wb = load_workbook(path, read_only=read_only)
        return wb
    except InvalidFileException:
        # traceback.print_exc()
        return None


# 获取所有sheet的名称
def get_sheet_names(wb):
    return wb.get_sheet_names()


# 获取sheet
def get_sheet_by_name(wb, name='Sheet1'):
    return wb.get_sheet_by_name(name)


# 读取sheet中的内容
def read(ws, row_s=-1, row_e=-1, col_s=-1, col_e=-1):
    r_s = 0 if row_s == -1 else row_s  # 起始行index
    r_e = ws.max_row if row_e == -1 else row_e  # 结束行index
    c_s = 0 if col_s == -1 else col_s  # 起始列index
    c_e = ws.max_column if col_e == -1 else col_e  # 结束列index

    rows = []
    for r_index in range(r_s, r_e):
        cols = []
        for c_index in range(c_s, c_e):
            value = ws.cell(row=r_index + 1, column=c_index + 1).value
            cols.append(value if value else '')
        rows.append(cols)

    return rows


# 读取sheet中某一列的数据
def read_col(ws, col_index, row_s=-1, row_e=-1):
    r_s = 0 if row_s == -1 else row_s  # 起始行index
    r_e = ws.max_row if row_e == -1 else row_e  # 结束行index

    rows = []
    for r_index in range(r_s, r_e):
        value = ws.cell(row=r_index+1, column=col_index+1).value
        rows.append(value if value else '')

    return rows


# 读取sheet中某一行的数据
def read_row(ws, row_index, col_s=-1, col_e=-1):
    c_s = 0 if col_s == -1 else col_s  # 起始列index
    c_e = ws.max_column if col_e == -1 else col_e  # 结束列index

    cols = []
    for c_index in range(c_s, c_e):
        value = ws.cell(row=row_index+1, column=c_index+1).value
        cols.append(value if value else '')

    return cols


# 读取sheet中某个cell的数据
def read_cell(ws, row_index=0, col_index=0):
    value = ws.cell(row=row_index+1, column=col_index+1).value
    return value if value else ''

