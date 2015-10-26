# coding=UTF-8
import traceback
import xlrd
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
import openpyxl

__author__ = 'jubileus'


'''
此工具类生成的excel在windows平台下无法打开
'''


# 打开xls文件
def open_excel(file='file.xls', encoding='utf-8'):
    try:
        xlrd.Book.encoding = encoding
        return xlrd.open_workbook(file)
    except Exception:
        traceback.print_exc()
        return None


# 获取所有sheet的名称
def get_sheet_names(wb):
    return wb.sheet_names()


# 根据sheet的index获取sheet
def get_sheet_by_index(wb, index=0):
    return wb.sheet_by_index(index)


# 根据sheet的name获取sheet
def get_sheet_by_name(wb, name='Sheet1'):
    return wb.sheet_by_name(name)


# 读取sheet中的内容
def read(ws, row_s=-1, row_e=-1, col_s=-1, col_e=-1):
    r_s = 0 if row_s == -1 else row_s  # 起始行index
    r_e = ws.nrows if row_e == -1 else row_e  # 结束行index
    c_s = 0 if col_s == -1 else col_s  # 起始列index
    c_e = ws.ncols if col_e == -1 else col_e  # 结束列index

    rows = []
    for r_index in range(r_s, r_e):
        row = ws.row_values(r_index)
        cols = []
        if row:
            for c_index in range(c_s, c_e):
                cols.append(row[c_index] if row[c_index] else '')
        else:
            for c_index in range(c_s, c_e):
                cols.append('')
        rows.append(cols)

    return rows


# 读取sheet中某一列的数据
def read_col(ws, col_index, row_s=-1, row_e=-1):
    r_s = 0 if row_s == -1 else row_s  # 起始行index
    r_e = ws.nrows if row_e == -1 else row_e  # 结束行index

    rows = []
    for r_index in range(r_s, r_e):
        row = ws.row_values(r_index)
        rows.append(row[col_index] if row and row[col_index] else '')

    return rows


# 读取sheet中某一行的数据
def read_row(ws, row_index, col_s=-1, col_e=-1):
    c_s = 0 if col_s == -1 else col_s  # 起始列index
    c_e = ws.ncols if col_e == -1 else col_e  # 结束列index

    row = ws.row_values(row_index)

    cols = []
    for c_index in range(c_s, c_e):
        if row:
            col = row[c_index]
            cols.append(col if col else '')
        else:
            cols.append('')

    return cols


# 读取sheet中某个cell的数据
def read_cell(ws, row_index=0, col_index=0):
    row = ws.row_values(row_index)
    return row[col_index] if row and row[col_index] else ''


# 创建要写入数据的workbook
def create_workbook():
    openpyxl.Workbook.encoding = 'gbk'
    return Workbook()


# 创建Writer
def create_writer(wb):
    return ExcelWriter(workbook=wb)


# 创建sheet
def create_sheet(wb, ws_name='Sheet1', ws_index=0):
    ws = wb.worksheets[ws_index]
    ws.title = ws_name
    return ws


# 保存workbook到磁盘
def save_workbook(ew, path='excel_1.xls'):
    ew.save(filename=path)
    
    
# 写入数据
def write(data, ws):
    for r in range(len(data)):
        row = data[r]
        for c in range(len(row)):
            ws.cell(row=r, column=c+1).value = row[c]
